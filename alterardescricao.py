import sys
import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import json

def main(usuario, senha, campos_selecionados, excel_path):
    executar_selenium(usuario, senha, campos_selecionados, excel_path)

def executar_selenium(usuario, senha, campos_selecionados, excel_path):
    relatorio = {
        "campos_selecionados": campos_selecionados,
        "total_processado": 0,
        "alterados": 0,
        "falhas": 0,
        "linhas_com_erro": []
    }

    try:
        df = pd.read_excel(excel_path)
        required_columns = ['cod_empresa', 'cod_item'] + campos_selecionados
        if not all(col in df.columns for col in required_columns):
            print(f"Erro: O Excel deve conter as colunas: {', '.join(required_columns)}")
            return
    except FileNotFoundError:
        print(f"Erro: Arquivo não encontrado: {excel_path}")
        return
    except Exception as e:
        print(f"Erro ao ler Excel: {str(e)}")
        return

    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(options=options)
    
    try:
        driver.get('http://192.168.0.6/')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'usuario')))

        # Login
        driver.find_element(By.NAME, 'usuario').send_keys(usuario)
        driver.find_element(By.NAME, 'senha').send_keys(senha + Keys.RETURN)
        time.sleep(1)
        driver.get('http://192.168.0.6/pcp/man_item.php')

        linhas_nao_alteradas = []
        itens_alterados = 0

        for index, row in df.iterrows():
            alterado = False
            try:
                # Preencher campos de busca
                driver.find_element(By.NAME, 'cod_empresa').clear()
                driver.find_element(By.NAME, 'cod_empresa').send_keys(str(row['cod_empresa']) + Keys.TAB)
                driver.find_element(By.NAME, 'cod_item').clear()
                driver.find_element(By.NAME, 'cod_item').send_keys(str(row['cod_item']) + Keys.RETURN)

                # Aguardar carregamento
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, 'den_item')))

                # Processar apenas campos selecionados
                campos = {campo: row[campo] for campo in campos_selecionados}

                for campo_nome, valor in campos.items():
                    try:
                        elemento = driver.find_element(By.NAME, campo_nome)
                        current_value = elemento.get_attribute('value')
                        
                        # Tratar valores NaN e strings vazias
                        desired_value = '' if pd.isna(valor) else str(valor).strip()
                        
                        if current_value == desired_value:
                            print(f"Campo {campo_nome} já está correto. Pulando.")
                            continue
                            
                        # Processar alteração
                        elemento.clear()
                        if desired_value:
                            elemento.send_keys(desired_value)
                        elemento.send_keys(Keys.RETURN)  # Confirmação IMEDIATA
                        
                        # Tratar alerta para cada campo
                        try:
                            WebDriverWait(driver, 2).until(EC.alert_is_present())
                            alerta = driver.switch_to.alert
                            print(f"Alerta em {campo_nome}: {alerta.text}")
                            alerta.accept()
                            alterado = True
                            time.sleep(0.3)  # Estabilização
                        except TimeoutException:
                            print(f"Nenhum alerta em {campo_nome}")
                            alterado = True

                    except Exception as e:
                        print(f"Erro em {campo_nome}: {str(e)}")
                        continue

                # Atualizar contadores
                    if alterado:
                        relatorio["alterados"] += 1
                    else:
                        relatorio["falhas"] += 1
                        linhas_nao_alteradas.append(index + 2)
                
                relatorio["total_processado"] = len(df)
                relatorio["linhas_com_erro"] = linhas_nao_alteradas
                
                relatorio_path = excel_path.replace('.xlsx', '_relatorio.json')
                with open(relatorio_path, 'w') as f:
                    json.dump(relatorio, f)
                    
                print(json.dumps(relatorio))

            except Exception as e:
                print(f"Erro no item {index + 1}: {str(e)}")
                linhas_nao_alteradas.append(index + 2)

                with open(relatorio_path, 'w') as f:
                    json.dump(relatorio, f)
                print(json.dumps(relatorio))

            # Reiniciar para próxima entrada
            driver.get('http://192.168.0.6/pcp/man_item.php')
            time.sleep(0.5)

        driver.quit()

        # Marcar células não alteradas
        if linhas_nao_alteradas:
            wb = load_workbook(excel_path)
            ws = wb.active
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for row in linhas_nao_alteradas:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = red_fill
            wb.save(excel_path)

        print(f"Relatório:\n"
              f"Campos selecionados: {', '.join(campos_selecionados)}\n"
              f"Total processado: {len(df)}\n"
              f"Alterados: {itens_alterados}\n"
              f"Falhas: {len(linhas_nao_alteradas)}\n"
              f"Linhas com erro: {linhas_nao_alteradas}")

    except Exception as e:
        if driver:
            driver.quit()
        print(f"Erro Fatal: Ocorreu um erro: {str(e)}")

if __name__ == "__main__":
    if len(sys.argv) < 5:
        print("Uso: python alterardescricao.py <usuario> <senha> <campos> <excel_path>")
        sys.exit(1)
    
    usuario = sys.argv[1]
    senha = sys.argv[2]
    campos_selecionados = sys.argv[3].split(',')
    excel_path = sys.argv[4]
    
    executar_selenium(usuario, senha, campos_selecionados, excel_path)